import json
import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from tools.standard_checklist_html import (
    ChecklistInputError,
    extract_checklist,
    load_answers,
    render_html,
)


class StandardChecklistHtmlTest(unittest.TestCase):
    def write_workbook(self, root):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "消防安全設備設置標準"
        sheet["A1"] = "14條下列場所應設置滅火器"
        sheet["B2"] = "一. 甲類場所,地下建築物,幼兒園."
        sheet["B3"] = "二. 總樓地板面積在一百五十平方公尺以上之乙,丙,丁類場所."
        sheet["A4"] = "15條下列場所應設置室內消防栓設備"
        sheet["B5"] = "一. 五層以下建築物,供第十二條第一款第一目所列場所使用."
        sheet["A1"].fill = PatternFill(fill_type="solid", fgColor="CCFFFF")
        sheet["A1"].font = Font(name="標楷體", size=12)
        sheet["B2"].alignment = Alignment(wrap_text=True, vertical="center")
        sheet.column_dimensions["A"].width = 5.625
        sheet.column_dimensions["B"].width = 70.625
        path = root / "standard.xlsx"
        workbook.save(path)
        return path

    def test_extracts_stable_answer_ids_from_article_item_rows(self):
        with tempfile.TemporaryDirectory() as tmp:
            checklist = extract_checklist(self.write_workbook(Path(tmp)))

            self.assertEqual([item.item_id for item in checklist.items], ["14-1", "14-2", "15-1"])
            self.assertEqual(checklist.items[0].row, 2)
            self.assertEqual(checklist.items[0].check_cell, "A2")
            self.assertIn("甲類場所", checklist.items[0].text)

    def test_renders_red_check_in_matching_check_cell_only(self):
        with tempfile.TemporaryDirectory() as tmp:
            checklist = extract_checklist(self.write_workbook(Path(tmp)))
            html = render_html(checklist, {"14-2"}, case_name="示範案件")

            self.assertIn("示範案件", html)
            self.assertIn('data-answer-id="14-2"', html)
            self.assertIn('<span class="red-check">✓</span>', html)
            self.assertNotIn('data-answer-id="14-1"><span class="red-check">✓</span>', html)

    def test_load_answers_rejects_unknown_ids_to_prevent_wrong_cells(self):
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            checklist = extract_checklist(self.write_workbook(root))
            answers_path = root / "answers.json"
            answers_path.write_text(
                json.dumps({"checked": ["14-1", "99-1"]}, ensure_ascii=False),
                encoding="utf-8",
            )

            with self.assertRaisesRegex(ChecklistInputError, "未知的檢核項目 ID: 99-1"):
                load_answers(answers_path, checklist)


if __name__ == "__main__":
    unittest.main()
