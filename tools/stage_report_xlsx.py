#!/usr/bin/env python3
"""兩階段審查 Excel 交付物產生工具 for fire-review.

第一階段：依 case.json 產出「複合用途建築物及樓層屬性檢討」工作簿（4 分頁）。
第二階段：依法條判斷表 xlsx ＋ 人工定案的 decisions.json 產出「設置標準檢討」工作簿（3 分頁）。

格式權威為 input/範例/ 內的兩份格式範本；本工具以程式複製其版面（分頁名稱、欄位順序、
合併儲存格、清單驗證、欄寬），不從空白表新造版面。

**本工具不做任何法規判斷。** 第二階段的勾選、備註、應設設備與待釐清事項一律來自
decisions.json——即經人工確認的判斷結果。工具只負責排版與格式一致性（最高原則 4、5）。

依賴 openpyxl（見 requirements.txt；bash tools/setup.sh 一鍵安裝）。

Usage:
    python3 tools/stage_report_xlsx.py first-stage --case output/case.json
    python3 tools/stage_report_xlsx.py stage-two \\
        --law rules/checklists/各類場所消防安全設備設置標準14~31條判斷用.xlsx \\
        --decisions output/stage2_decisions.json \\
        --case output/case.json
"""

import argparse
import json
import os
import re
import sys

try:
    from tools.console import force_utf8_output
except ImportError:  # 直接以 tools/ 為工作目錄執行時
    from console import force_utf8_output

force_utf8_output()

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
except ImportError:  # pragma: no cover - 環境未安裝時的明確指引
    sys.exit("缺少 openpyxl。請執行：bash tools/setup.sh   或   python3 -m pip install -r requirements.txt")

MANUAL = "⚪需人工判讀"
OUT_OF_SCOPE = "非本次申請範圍"

DEFAULT_LAW_PATH = os.path.join("rules", "checklists", "各類場所消防安全設備設置標準14~31條判斷用.xlsx")
STAGE_TWO_SHEET = "消防安全設備設置標準"

THIN = Side(style="thin", color="999999")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HEADER_FILL = PatternFill("solid", fgColor="D9EAF7")
TOTAL_FILL = PatternFill("solid", fgColor="F2F2F2")
CHECK_FILL = PatternFill("solid", fgColor="E2F0D9")
HEADER_FONT = Font(bold=True, color="17365D")
RED_FONT = Font(color="FF0000")
TITLE_FONT = Font(bold=True, size=14)
WRAP_TOP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)

ARTICLE_18_OPTIONS_PATH = os.path.join("rules", "article18_equipment_options.json")
ITEM_NUMERALS = {1: "一", 2: "二", 3: "三", 4: "四", 5: "五",
                 6: "六", 7: "七", 8: "八", 9: "九", 10: "十"}

EQUIPMENT_BY_ARTICLE = {
    "14": "滅火器", "15": "室內消防栓設備", "16": "室外消防栓設備", "17": "自動撒水設備",
    "18": "自動滅火設備", "19": "火警自動警報設備", "20": "手動報警設備",
    "21": "瓦斯漏氣火警自動警報設備", "22": "緊急廣播設備", "22-1": "一一九火災通報裝置",
    "23": "標示設備", "24": "緊急照明設備", "25": "避難器具", "26": "連結送水管",
    "27": "消防專用蓄水池", "28": "排煙設備", "29": "緊急電源插座",
    "30": "無線電通信輔助設備", "31": "防災監控系統綜合操作裝置",
}

CHINESE_ITEM = {"一": 1, "二": 2, "三": 3, "四": 4, "五": 5,
                "六": 6, "七": 7, "八": 8, "九": 9, "十": 10}


# ---------------------------------------------------------------- 共用小工具

def get_value(field):
    if isinstance(field, dict) and "value" in field:
        return field["value"]
    return field


def get_confidence(field):
    return field.get("confidence") if isinstance(field, dict) else None


def set_widths(sheet, widths):
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width


def write_header(sheet, row, headers):
    for column, text in enumerate(headers, start=1):
        cell = sheet.cell(row=row, column=column, value=text)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = BORDER
        cell.alignment = CENTER


def style_body(sheet, first_row, last_row, column_count):
    for row in range(first_row, last_row + 1):
        for column in range(1, column_count + 1):
            cell = sheet.cell(row=row, column=column)
            cell.border = BORDER
            cell.alignment = WRAP_TOP


# ------------------------------------------------------------ 第一階段工作簿

def floor_attribute(floor):
    """樓層屬性；證據不足一律留白（刻意保留的空白選項，不得以推測填充）。"""
    position = get_value(floor.get("floor_position"))
    if get_confidence(floor.get("floor_position")) == "low" or position is None:
        return ""
    if position == "basement":
        return "地下層"
    if position in ("roof_level", "roof"):
        return "屋頂層"
    if get_value(floor.get("windowless")) is True:
        return "無開口樓層"
    if get_value(floor.get("windowless")) is None:
        return ""
    return "一般樓層"


def use_text(floor):
    use = floor.get("use_category") or {}
    if get_confidence(use) == "low":
        return MANUAL
    label = use.get("label") if isinstance(use, dict) else None
    value = get_value(use)
    if not value and not label:
        return MANUAL
    return f"{label}（{value}）" if label and value else str(label or value)


def classification_lines(case):
    """§12 分類逐項列出（結構化，一項一列），不合併成散文。

    每一項分類都必須明示第 12 條的款、目及類別（rules/review_corrections.md
    2026-07-13）。case.json 未載 legal_basis 者不得靜默略過——標記為需人工判讀，
    以維持法條可追溯（審圖最高原則 1）。
    """
    seen = {}
    for floor in case.get("floors") or []:
        use = floor.get("use_category") or {}
        value = get_value(use)
        if not value:
            continue
        legal = use.get("legal_basis") if isinstance(use, dict) else None
        label = use.get("label") if isinstance(use, dict) else None
        if legal:
            text = f"{legal}（{value}）{label}" if label else f"{legal}（{value}）"
        else:
            text = f"{value}{'　' + label if label else ''}　{MANUAL}：§12 款目未載於 case.json"
        seen.setdefault(text, None)
    return list(seen) or [MANUAL]


def judgment_text(case):
    assessment = ((case.get("building") or {}).get("mixed_use_assessment")) or {}
    if assessment.get("source") != "manual" or assessment.get("is_mixed_use") is None:
        return ""
    if assessment.get("is_mixed_use") is False:
        return "單一用途建築物"
    candidate = str(assessment.get("category_candidate") or "")
    if "戊1" in candidate or "第5款第1目" in candidate or "戊類第一目" in candidate:
        return "戊1複合用途建築物"
    if "戊2" in candidate or "第5款第2目" in candidate or "戊類第二目" in candidate:
        return "戊2複合用途建築物"
    return ""


def build_first_stage(case):
    workbook = Workbook()
    main = workbook.active
    main.title = "複合用途檢討"

    permit_no = get_value(((case.get("use_permit") or {}).get("permit_no"))) or MANUAL
    renovation = case.get("interior_renovation") or {}
    renovation_floors = get_value(renovation.get("floors")) or []
    renovation_area = get_value(renovation.get("area"))
    post_use = get_value(renovation.get("post_renovation_use"))

    main["A1"] = "複合用途建築物及樓層屬性檢討"
    main["A1"].font = TITLE_FONT
    main.merge_cells("A1:G1")
    main["A2"] = (f"案名：{case.get('case_name') or MANUAL}"
                  f" / 使用執照：{permit_no}"
                  f" / 本次申請範圍：{renovation_area if renovation_area is not None else MANUAL} m²")
    main.merge_cells("A2:G2")
    main["A2"].alignment = WRAP_TOP

    write_header(main, 4, ["樓層", "各層用途", "樓地板面積(m²)", "各層變更後用途",
                           "本次申請範圍樓地板面積(m²)", "樓層屬性", "備註"])

    first_row = 5
    row = first_row
    for floor in case.get("floors") or []:
        label = floor.get("floor") or MANUAL
        in_scope = label in renovation_floors
        area = get_value(floor.get("area"))
        main.cell(row=row, column=1, value=label)
        main.cell(row=row, column=2, value=use_text(floor))
        main.cell(row=row, column=3, value=area if area is not None else MANUAL)
        main.cell(row=row, column=4, value=post_use if in_scope and post_use else None)
        main.cell(row=row, column=5, value=renovation_area if in_scope and renovation_area is not None else None)
        main.cell(row=row, column=6, value=floor_attribute(floor))
        main.cell(row=row, column=7, value="本次申請範圍" if in_scope else None)
        row += 1

    # 最後一筆實際樓層後保留三列空白（rules/review_corrections.md 2026-07-15）
    blank_rows = 3
    last_body_row = row + blank_rows - 1
    style_body(main, first_row, last_body_row, 7)

    total_row = last_body_row + 1
    main.cell(row=total_row, column=1, value="樓地板面積合計").font = Font(bold=True)
    main.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=2)
    main.cell(row=total_row, column=3, value=f"=SUM(C{first_row}:C{last_body_row})")
    main.cell(row=total_row, column=5, value=f"=SUM(E{first_row}:E{last_body_row})")
    for column in range(1, 8):
        cell = main.cell(row=total_row, column=column)
        cell.border = BORDER
        cell.fill = TOTAL_FILL
        cell.font = Font(bold=True)

    # 合計列後保留一列空白間隔，再進判定區
    verdict_row = total_row + 2
    main.cell(row=verdict_row, column=1, value="複合用途建築物判定").font = HEADER_FONT
    main.merge_cells(start_row=verdict_row, start_column=1, end_row=verdict_row, end_column=7)
    main.cell(row=verdict_row, column=1).fill = HEADER_FILL

    write_header(main, verdict_row + 1, ["第12條用途分類", None, None, None, "本場所供判定", None, None])
    main.merge_cells(start_row=verdict_row + 1, start_column=1, end_row=verdict_row + 1, end_column=4)
    main.merge_cells(start_row=verdict_row + 1, start_column=5, end_row=verdict_row + 1, end_column=7)

    lines = classification_lines(case)
    verdict = judgment_text(case)
    class_first = verdict_row + 2
    for offset, text in enumerate(lines):
        target = class_first + offset
        main.cell(row=target, column=1, value=text).alignment = WRAP_TOP
        main.merge_cells(start_row=target, start_column=1, end_row=target, end_column=4)
        # 判定只在第一列適用列填一次
        main.cell(row=target, column=5, value=verdict if offset == 0 else None).alignment = WRAP_TOP
        main.merge_cells(start_row=target, start_column=5, end_row=target, end_column=7)
        for column in (1, 5):
            main.cell(row=target, column=column).border = BORDER
    class_last = class_first + max(len(lines), 1) - 1

    attribute_dv = DataValidation(type="list", formula1='"地下層,無開口樓層,屋頂層,一般樓層,"', allow_blank=True)
    scope_dv = DataValidation(type="list", formula1='"本次申請範圍,"', allow_blank=True)
    verdict_dv = DataValidation(
        type="list", formula1='"單一用途建築物,戊1複合用途建築物,戊2複合用途建築物,"', allow_blank=True)
    main.add_data_validation(attribute_dv)
    main.add_data_validation(scope_dv)
    main.add_data_validation(verdict_dv)
    attribute_dv.add(f"F{first_row}:F{last_body_row}")
    scope_dv.add(f"G{first_row}:G{last_body_row}")
    verdict_dv.add(f"E{class_first}:E{class_last}")

    set_widths(main, [14, 24, 16, 24, 21, 15, 28])
    main.freeze_panes = "A5"

    basis = workbook.create_sheet("用途判斷依據")
    write_header(basis, 1, ["用途", "對應樓層", "來源文字", "第十二條分類", "判斷說明"])
    row = 2
    for floor in case.get("floors") or []:
        use = floor.get("use_category") or {}
        relation = floor.get("use_relation") or {}
        basis.cell(row=row, column=1, value=use.get("label") if isinstance(use, dict) else None)
        basis.cell(row=row, column=2, value=floor.get("floor"))
        basis.cell(row=row, column=3, value=(use.get("source") if isinstance(use, dict) else None))
        basis.cell(row=row, column=4, value=(use.get("legal_basis") if isinstance(use, dict) else None))
        basis.cell(row=row, column=5, value=relation.get("basis"))
        row += 1
    style_body(basis, 2, max(row - 1, 2), 5)
    set_widths(basis, [18, 34, 36, 22, 58])

    sources = workbook.create_sheet("來源資料")
    write_header(sources, 1, ["項目", "擷取內容", "來源"])
    permit = case.get("use_permit") or {}
    rows = [
        ("使用執照字號", permit_no, (permit.get("permit_no") or {}).get("source")),
        ("核發日期", get_value(permit.get("issued_date")), (permit.get("issued_date") or {}).get("source")),
        ("原核准總樓地板面積", get_value(permit.get("total_floor_area")),
         (permit.get("total_floor_area") or {}).get("source")),
        ("本次申請樓層", "、".join(str(f) for f in renovation_floors) or MANUAL,
         (renovation.get("floors") or {}).get("source")),
        ("本次申請範圍樓地板面積", renovation_area, (renovation.get("area") or {}).get("source")),
        ("工程類別", get_value(renovation.get("works_type")), (renovation.get("works_type") or {}).get("source")),
        ("法規版本", case.get("regulation_version"), "rules/*.json"),
    ]
    for offset, (item, content, source) in enumerate(rows):
        sources.cell(row=2 + offset, column=1, value=item)
        sources.cell(row=2 + offset, column=2, value=content if content is not None else MANUAL)
        sources.cell(row=2 + offset, column=3, value=source or MANUAL)
    style_body(sources, 2, 1 + len(rows), 3)
    set_widths(sources, [34, 52, 42])

    heights = workbook.create_sheet("各樓層高度")
    heights["A1"] = "各樓層高度"
    heights["A1"].font = TITLE_FONT
    heights.merge_cells("A1:D1")
    write_header(heights, 3, ["樓層", "樓層高度(m)", "資料來源", "備註"])
    row = 4
    for floor in case.get("floors") or []:
        rooms = floor.get("rooms") or []
        height = next((r.get("ceiling_height") for r in rooms if r.get("ceiling_height") is not None), None)
        heights.cell(row=row, column=1, value=floor.get("floor"))
        heights.cell(row=row, column=2, value=height)
        heights.cell(row=row, column=3, value="case.json rooms[].ceiling_height" if height is not None else None)
        heights.cell(row=row, column=4, value=None if height is not None else "未提供")
        row += 1
    style_body(heights, 4, max(row - 1, 4), 4)
    set_widths(heights, [18, 18, 34, 30])

    return workbook


# ------------------------------------------------------------ 第二階段工作簿

def normalize_law_rows(values):
    """法條判斷表 → [(條文/設備, 款項條件)]。相容原始三欄與已含輸出表頭的四欄格式。"""
    def text(value):
        stripped = str(value).strip() if value is not None else ""
        return stripped or None

    rows = list(values)
    has_output_header = (rows and text(rows[0][0]) == "判斷"
                         and len(rows[0]) > 1 and text(rows[0][1]) == "條文/設備")
    if has_output_header:
        rows = rows[1:]
    normalized = []
    for row in rows:
        padded = list(row) + [None] * 4
        title, condition = (padded[1], padded[2]) if has_output_header else (padded[0], padded[1])
        if text(title) or text(condition):
            normalized.append((text(title), text(condition)))
    return normalized


def article_18_fallback(path=None):
    """§18 表列各款 → [(條文/設備, 款項條件)]。

    唯一來源是 rules/article18_equipment_options.json——條文轉錄只存在一份，
    避免多處複本隨修法漂移（見 rules/stage_two_judgment_rules.md「維護規則」）。
    """
    doc = load_article18_doc(path)
    if not doc:
        return []
    rows = [(f"18條{doc.get('article_title', '下列場所應設置自動滅火設備')}", None)]
    for entry in doc.get("items") or []:
        numeral = ITEM_NUMERALS.get(entry.get("item"), str(entry.get("item")))
        rows.append((None, f"{numeral}. {entry.get('condition', '')}"))
    return rows


def ensure_article_18(rows, path=None):
    """§18 必須出現在 §17 與 §19 之間；法條判斷表缺漏時自規則檔補入。"""
    if any(title and title.startswith("18條") for title, _ in rows):
        return rows
    fallback = article_18_fallback(path)
    if not fallback:
        return rows
    insert_at = next((i for i, (title, _) in enumerate(rows) if title and title.startswith("19條")), None)
    if insert_at is None:
        return rows + fallback
    return rows[:insert_at] + fallback + rows[insert_at:]


def parse_article(title):
    match = re.match(r"^(\d+(?:-\d+)?)條", title or "")
    return match.group(1) if match else None


def parse_item(condition):
    match = re.match(r"^([一二三四五六七八九十])[.、．]", condition or "")
    return CHINESE_ITEM.get(match.group(1)) if match else None


def add_references(rows):
    """為每列標上條款代號（例：14-2、22）。純標題列且後續有同條款項者不給代號。"""
    referenced = []
    current_article = None
    for title, condition in rows:
        article = parse_article(title) or current_article
        if parse_article(title):
            current_article = parse_article(title)
        item = parse_item(condition)
        reference = None
        if article:
            reference = f"{article}-{item}" if item else article
        referenced.append({"title": title, "condition": condition,
                           "article": article, "item": item, "reference": reference})
    for index, row in enumerate(referenced):
        following = referenced[index + 1] if index + 1 < len(referenced) else None
        if row["title"] and not row["condition"] and following \
                and following["article"] == row["article"] and following["condition"]:
            row["reference"] = None
    return referenced


def sort_refs(refs):
    def key(ref):
        article, _, item = str(ref).partition("-")
        return (float(article.replace("-", ".")) if article.replace(".", "").isdigit() else 0.0,
                int(item) if item.isdigit() else 0)
    return sorted(refs, key=key)


def format_basis(refs):
    by_article = {}
    for ref in sort_refs(refs):
        article, _, item = str(ref).partition("-")
        by_article.setdefault(article, [])
        if item:
            by_article[article].append(item)
    parts = []
    for article, items in by_article.items():
        parts.append(f"第{article}條第{'、'.join(items)}款" if items else f"第{article}條")
    return "、".join(parts)


def scope_only_note(note):
    """備註僅可記載 `非本次申請範圍`；其餘一律留白（rules/review_corrections.md 2026-07-15）。"""
    return OUT_OF_SCOPE if OUT_OF_SCOPE in str(note or "") else ""


def default_pending_rows():
    return [["無", "目前依第一階段資料可完成本次第二階段判斷。", "", ""]]


def build_equipment_rows(decisions):
    rows = decisions.get("equipmentRows")
    if rows:
        return rows
    by_article = {}
    for ref in decisions.get("checkedRefs") or []:
        article = str(ref).partition("-")[0]
        if article not in EQUIPMENT_BY_ARTICLE:
            continue
        by_article.setdefault(article, []).append(ref)
    return [[EQUIPMENT_BY_ARTICLE[article], format_basis(refs), ""]
            for article, refs in by_article.items()]


def build_stage_two(law_rows, decisions):
    checked = {str(ref) for ref in (decisions.get("checkedRefs") or [])}
    notes = {str(k): v for k, v in (decisions.get("notesByRef") or {}).items()}

    rows = add_references(ensure_article_18(law_rows))
    # §18 為選擇設置，「條文/設備」欄僅列該款可選設備；第 8、9 款不是設備觸發條件
    options = load_article18_options()
    rows = [r for r in rows if not (r["article"] == "18" and r["item"] and r["item"] > 7)]
    for row in rows:
        if row["article"] == "18" and not row["item"]:
            row["title"] = "18條下列場所應設置自動滅火設備"
        elif row["article"] == "18" and row["item"] in options:
            row["title"] = options[row["item"]]

    workbook = Workbook()
    main = workbook.active
    main.title = STAGE_TWO_SHEET
    write_header(main, 1, ["判斷", "條文/設備", "款項條件", "備註"])
    for offset, row in enumerate(rows):
        target = 2 + offset
        mark = "✓" if row["reference"] and row["reference"] in checked else ""
        note = scope_only_note(notes.get(row["reference"])) if row["reference"] else ""
        main.cell(row=target, column=1, value=mark)
        main.cell(row=target, column=2, value=row["title"])
        main.cell(row=target, column=3, value=row["condition"])
        cell = main.cell(row=target, column=4, value=note)
        if note:
            cell.font = RED_FONT
    style_body(main, 2, 1 + len(rows), 4)
    for offset in range(len(rows)):
        cell = main.cell(row=2 + offset, column=1)
        cell.alignment = CENTER
        cell.fill = CHECK_FILL
        cell.font = Font(bold=True, color="375623")
    set_widths(main, [10, 36, 110, 42])
    main.freeze_panes = "A2"

    equipment_rows = [[r[0] if len(r) > 0 else "", r[1] if len(r) > 1 else "",
                       scope_only_note(r[2] if len(r) > 2 else "")]
                      for r in build_equipment_rows(decisions)]
    equipment = workbook.create_sheet("應設置設備")
    write_header(equipment, 1, ["消防設備", "符合依據", "備註"])
    for offset, row in enumerate(equipment_rows):
        for column, value in enumerate(row, start=1):
            cell = equipment.cell(row=2 + offset, column=column, value=value)
            if column == 3 and value:
                cell.font = RED_FONT
    style_body(equipment, 2, 1 + max(len(equipment_rows), 1), 3)
    set_widths(equipment, [28, 80, 42])
    equipment.freeze_panes = "A2"

    pending_rows = decisions.get("pendingRows") or default_pending_rows()
    pending = workbook.create_sheet("待釐清事項")
    write_header(pending, 1, ["條文", "疑問", "目前缺少資料", "影響"])
    for offset, row in enumerate(pending_rows):
        padded = list(row) + [""] * 4
        for column in range(1, 5):
            pending.cell(row=2 + offset, column=column, value=padded[column - 1] or "")
    style_body(pending, 2, 1 + max(len(pending_rows), 1), 4)
    set_widths(pending, [18, 60, 70, 50])
    pending.freeze_panes = "A2"

    return workbook, rows


def load_article18_doc(path=None):
    """§18 規則檔內容；不存在時回傳 None（呼叫端各自降級）。"""
    path = path or ARTICLE_18_OPTIONS_PATH
    if not os.path.exists(path):
        return None
    with open(path, encoding="utf-8") as f:
        return json.load(f)


def load_article18_options(path=None):
    """§18 各款可選設備 → {款次: "泡沫、乾粉"}。規則檔缺漏時回傳空 dict。"""
    doc = load_article18_doc(path)
    if not doc:
        return {}
    options = {}
    for entry in doc.get("items") or []:
        names = list(entry.get("options") or [])
        if not names:
            continue
        text = "、".join(names)
        if entry.get("alternative_sprinkler"):
            text += f"；或{entry['alternative_sprinkler']}"
        options[entry["item"]] = text
    return options


# ---------------------------------------------------------------------- CLI

def read_law_rows(path):
    workbook = load_workbook(path, data_only=True)
    sheet = workbook[STAGE_TWO_SHEET] if STAGE_TWO_SHEET in workbook.sheetnames else workbook.worksheets[0]
    return normalize_law_rows(sheet.iter_rows(values_only=True))


def default_output(case_path, case, suffix):
    name = (case or {}).get("case_name") or "案件"
    return os.path.join(os.path.dirname(case_path or "."), f"{name}-{suffix}.xlsx")


def cmd_first_stage(args):
    with open(args.case, encoding="utf-8") as f:
        case = json.load(f)
    out = args.output or default_output(args.case, case, "第一階段-複合用途及樓層屬性檢討")
    build_first_stage(case).save(out)
    floors = len(case.get("floors") or [])
    print(f"✅ 已輸出第一階段工作簿：{out}（{floors} 層；4 分頁）")
    print("   格式權威：input/範例/第一階段-複合用途及樓層屬性檢討-格式範本.xlsx")
    print("   樓層屬性留白為刻意保留的未分類狀態，嚴禁以推測填充。")
    return 0


def cmd_stage_two(args):
    case = None
    if args.case:
        with open(args.case, encoding="utf-8") as f:
            case = json.load(f)
    decisions = {}
    if args.decisions:
        with open(args.decisions, encoding="utf-8") as f:
            decisions = json.load(f)

    law_rows = read_law_rows(args.law)
    workbook, rows = build_stage_two(law_rows, decisions)
    out = args.output or default_output(args.case, case, "第二階段-設置標準檢討")
    workbook.save(out)

    articles = {row["article"] for row in rows if row["article"]}
    checked = len(decisions.get("checkedRefs") or [])
    print(f"✅ 已輸出第二階段工作簿：{out}（{len(rows)} 列；勾選 {checked} 款；3 分頁）")
    if "18" not in articles:
        print("⚠️  第 18 條未出現在輸出中——請檢查法條判斷表。")
    if not decisions:
        print("⚠️  未提供 decisions.json：全表未勾選。勾選結果必須由人工判斷後填入，工具不代為判定。")
    print("   判斷依據：rules/stage_two_judgment_rules.md（未經核定，須附警語）")
    return 0


def main(argv=None):
    parser = argparse.ArgumentParser(description="兩階段審查 Excel 交付物產生")
    sub = parser.add_subparsers(dest="command", required=True)

    first = sub.add_parser("first-stage", help="第一階段：複合用途建築物及樓層屬性檢討")
    first.add_argument("--case", required=True, help="case.json 路徑")
    first.add_argument("--output", help="輸出 xlsx 路徑")
    first.set_defaults(func=cmd_first_stage)

    second = sub.add_parser("stage-two", help="第二階段：設置標準檢討")
    second.add_argument("--law", default=DEFAULT_LAW_PATH, help="法條判斷表 xlsx 路徑")
    second.add_argument("--decisions", help="人工定案的 decisions.json（checkedRefs/notesByRef/equipmentRows/pendingRows）")
    second.add_argument("--case", help="case.json 路徑（供命名與輸出目錄）")
    second.add_argument("--output", help="輸出 xlsx 路徑")
    second.set_defaults(func=cmd_stage_two)

    args = parser.parse_args(argv)
    return args.func(args)


if __name__ == "__main__":
    sys.exit(main())
