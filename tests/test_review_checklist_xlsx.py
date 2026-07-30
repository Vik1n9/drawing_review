import contextlib
import io
import json
import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook

from tools.review_checklist_xlsx import (
    build_checklist,
    main,
    STATUS_MAP,
    RULE_STATUS_LABEL,
)


def sample_results():
    return {
        "case_name": "測試案",
        "date": "2026-07-31",
        "regulation_version": "測試版 v1",
        "regulation_html": "../rules/regulation-checklist.html",
        "article_range": "§14~§31（逐條窮舉）",
        "not_coded_articles": ["§26", "§27"],
        "items": [
            {"article": "§14", "anchor": "art-14", "equipment": "滅火器", "floor": "1F",
             "requirement": "甲類場所應設", "status": "pass",
             "finding": "符合：已設滅火器 4 具，足額", "rule_status": "coded",
             "source_page": None},
            {"article": "§15", "anchor": "art-15", "equipment": "室內消防栓", "floor": "B1",
             "requirement": "地下層應設", "status": "fail",
             "finding": "法定應設而圖面未見既有設備（重大缺失候選）", "rule_status": "coded",
             "source_page": None},
            {"article": "§19", "anchor": "art-19", "equipment": "火警自動警報設備", "floor": "3F",
             "requirement": "六層以上應設", "status": "manual",
             "finding": "需人工判讀：樓層屬性未確認｜本參數尚未逐條確認", "rule_status": "coded",
             "source_page": None},
            {"article": "§26", "anchor": "art-26", "equipment": "連結送水管", "floor": "—",
             "requirement": "（第 26 條 條文原文未索引）", "status": "manual",
             "finding": "⚪ 需人工判讀（規則未入庫）", "rule_status": "not_coded",
             "source_page": None},
            {"article": "§30", "anchor": "art-30", "equipment": "無線電通信輔助設備", "floor": "全棟",
             "requirement": "高層建築應設", "status": "na",
             "finding": "不適用：本案非高層建築", "rule_status": "coded",
             "source_page": None},
        ],
    }


class ChecklistWorkbookTest(unittest.TestCase):
    def setUp(self):
        self.workbook = build_checklist(sample_results())
        self.main = self.workbook["法條檢核清單"]

    def test_sheet_structure(self):
        self.assertEqual(self.workbook.sheetnames, ["法條檢核清單", "摘要統計"])

    def test_title_row_contains_case_name(self):
        self.assertIn("測試案", self.main["A1"].value)

    def test_meta_row_contains_date_and_version(self):
        meta = self.main["A2"].value
        self.assertIn("2026-07-31", meta)
        self.assertIn("測試版 v1", meta)

    def test_column_headers(self):
        headers = [self.main.cell(row=4, column=c).value for c in range(1, 9)]
        self.assertEqual(headers, ["#", "檢核狀態", "法條", "樓層", "設備", "應設要求", "檢核說明", "規則狀態"])

    def test_data_rows_match_items(self):
        results = sample_results()
        for idx, item in enumerate(results["items"]):
            row = 5 + idx
            self.assertEqual(self.main.cell(row=row, column=1).value, idx + 1)
            self.assertEqual(self.main.cell(row=row, column=3).value, item["article"])
            self.assertEqual(self.main.cell(row=row, column=5).value, item["equipment"])

    def test_status_fills_applied_correctly(self):
        # pass → 淺綠
        pass_fill = self.main.cell(row=5, column=2).fill
        self.assertEqual(pass_fill.fgColor.rgb, "00E2F0D9")
        # fail → 淺紅
        fail_fill = self.main.cell(row=6, column=2).fill
        self.assertEqual(fail_fill.fgColor.rgb, "00FDECEC")
        # manual → 淺灰
        manual_fill = self.main.cell(row=7, column=2).fill
        self.assertEqual(manual_fill.fgColor.rgb, "00F5F5F5")

    def test_rule_status_labels(self):
        self.assertEqual(self.main.cell(row=5, column=8).value, "已入庫（門檻）")
        self.assertEqual(self.main.cell(row=8, column=8).value, "未入庫")

    def test_summary_row_at_bottom(self):
        # 5 items → rows 5-9, summary at row 10
        summary_row = 10
        self.assertEqual(self.main.cell(row=summary_row, column=1).value, "摘要")
        summary_text = self.main.cell(row=summary_row, column=3).value
        self.assertIn("共 5 項", summary_text)
        self.assertIn("符合 1", summary_text)
        self.assertIn("不符合 1", summary_text)
        self.assertIn("需人工判讀 2", summary_text)
        self.assertIn("不適用 1", summary_text)

    def test_freeze_panes_set(self):
        self.assertEqual(self.main.freeze_panes, "A5")


class SummarySheetTest(unittest.TestCase):
    def setUp(self):
        self.summary = build_checklist(sample_results())["摘要統計"]

    def test_status_counts(self):
        # Row 4 = pass, Row 5 = fail, Row 6 = manual, Row 7 = na
        self.assertEqual(self.summary.cell(row=4, column=2).value, 1)  # pass
        self.assertEqual(self.summary.cell(row=5, column=2).value, 1)  # fail
        self.assertEqual(self.summary.cell(row=6, column=2).value, 2)  # manual
        self.assertEqual(self.summary.cell(row=7, column=2).value, 1)  # na

    def test_not_coded_articles_listed(self):
        # Row 10 contains the not-coded articles
        not_coded_text = self.summary.cell(row=10, column=1).value
        self.assertIn("§26", not_coded_text)
        self.assertIn("§27", not_coded_text)

    def test_unverified_parameter_count(self):
        # Row 13 contains unverified parameter stats
        unverified_text = self.summary.cell(row=13, column=1).value
        self.assertIn("1 / 5", unverified_text)


class CliTest(unittest.TestCase):
    def test_main_writes_xlsx(self):
        with tempfile.TemporaryDirectory() as tmp:
            results_path = Path(tmp) / "check_results.json"
            results_path.write_text(json.dumps(sample_results(), ensure_ascii=False), encoding="utf-8")

            with contextlib.redirect_stdout(io.StringIO()):
                self.assertEqual(main(["--results", str(results_path)]), 0)

            output_path = Path(tmp) / "測試案-法條檢核清單.xlsx"
            self.assertTrue(output_path.exists())
            wb = load_workbook(output_path)
            self.assertEqual(wb.sheetnames, ["法條檢核清單", "摘要統計"])

    def test_custom_output_path(self):
        with tempfile.TemporaryDirectory() as tmp:
            results_path = Path(tmp) / "check_results.json"
            results_path.write_text(json.dumps(sample_results(), ensure_ascii=False), encoding="utf-8")
            custom_out = Path(tmp) / "custom.xlsx"

            with contextlib.redirect_stdout(io.StringIO()):
                self.assertEqual(main(["--results", str(results_path), "--output", str(custom_out)]), 0)

            self.assertTrue(custom_out.exists())


if __name__ == "__main__":
    unittest.main()
