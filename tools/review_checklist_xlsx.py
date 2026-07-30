#!/usr/bin/env python3
"""法條檢核清單 Excel 產生工具 for fire-review.

依 check_results.json 產出「法條檢核清單」工作簿（2 分頁）：
- 分頁 1：逐條檢核明細（§14~§31），含狀態符號與底色
- 分頁 2：摘要統計（各狀態數量、規則未入庫條號、未核定參數比例）

格式與 stage_report_xlsx.py 保持一致（共用排版慣例）。

依賴 openpyxl（見 requirements.txt；bash tools/setup.sh 一鍵安裝）。

Usage:
    python3 tools/review_checklist_xlsx.py --results output/check_results.json
    python3 tools/review_checklist_xlsx.py --results output/check_results.json --output output/自訂.xlsx
"""

import argparse
import json
import os
import sys
from datetime import date

try:
    from tools.console import force_utf8_output
except ImportError:  # 直接以 tools/ 為工作目錄執行時
    from console import force_utf8_output

force_utf8_output()

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
except ImportError:  # pragma: no cover - 環境未安裝時的明確指引
    sys.exit("缺少 openpyxl。請執行：bash tools/setup.sh   或   python3 -m pip install -r requirements.txt")

# ---------------------------------------------------------------- 排版常數

THIN = Side(style="thin", color="999999")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HEADER_FILL = PatternFill("solid", fgColor="D9EAF7")
HEADER_FONT = Font(bold=True, color="17365D")
TITLE_FONT = Font(bold=True, size=14)
WRAP_TOP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)

PASS_FILL = PatternFill("solid", fgColor="E2F0D9")
FAIL_FILL = PatternFill("solid", fgColor="FDECEC")
MANUAL_FILL = PatternFill("solid", fgColor="F5F5F5")
NA_FILL = PatternFill("solid", fgColor="FFFFFF")
TOTAL_FILL = PatternFill("solid", fgColor="F2F2F2")

STATUS_MAP = {
    "pass":   ("☑ 符合", PASS_FILL),
    "fail":   ("☒ 不符合", FAIL_FILL),
    "manual": ("⚪ 需人工判讀", MANUAL_FILL),
    "na":     ("— 不適用", NA_FILL),
}

RULE_STATUS_LABEL = {
    "coded": "已入庫（門檻）",
    "coded_quantity": "已入庫（數量/配置）",
    "not_coded": "未入庫",
}


# ---------------------------------------------------------------- 共用小工具

def set_widths(sheet, widths):
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width


def write_header(sheet, row, headers):
    for column, text in enumerate(headers, start=1):
        cell = sheet.cell(row=row, column=column, value=text)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = BORDER
        cell.alignment = CENTER


def style_body(sheet, first_row, last_row, column_count):
    for row in range(first_row, last_row + 1):
        for column in range(1, column_count + 1):
            cell = sheet.cell(row=row, column=column)
            cell.border = BORDER
            cell.alignment = WRAP_TOP


# ---------------------------------------------------------------- 工作簿建構

def build_checklist(results):
    """依 check_results.json 產出法條檢核清單工作簿。"""
    workbook = Workbook()

    # ---- 分頁 1：法條檢核清單 ----
    main = workbook.active
    main.title = "法條檢核清單"

    case_name = results.get("case_name", "(未命名案件)")
    check_date = results.get("date", str(date.today()))
    reg_version = results.get("regulation_version", "未注明")
    article_range = results.get("article_range", "§14~§31")

    # 表頭資訊
    main["A1"] = f"法條檢核清單 — {case_name}"
    main["A1"].font = TITLE_FONT
    main.merge_cells("A1:H1")
    main["A2"] = (f"檢核日期：{check_date}｜法規版本：{reg_version}｜檢核範圍：{article_range}")
    main["A2"].alignment = WRAP_TOP
    main.merge_cells("A2:H2")

    # 欄位標題
    write_header(main, 4, ["#", "檢核狀態", "法條", "樓層", "設備", "應設要求", "檢核說明", "規則狀態"])

    # 逐筆資料
    first_row = 5
    row = first_row
    for idx, item in enumerate(results.get("items", []), start=1):
        status = item.get("status", "manual")
        label, fill = STATUS_MAP.get(status, STATUS_MAP["manual"])

        main.cell(row=row, column=1, value=idx)
        main.cell(row=row, column=1).alignment = CENTER

        status_cell = main.cell(row=row, column=2, value=label)
        status_cell.fill = fill
        status_cell.alignment = CENTER

        main.cell(row=row, column=3, value=item.get("article", ""))
        main.cell(row=row, column=4, value=str(item.get("floor", "—")))
        main.cell(row=row, column=5, value=item.get("equipment", ""))
        main.cell(row=row, column=6, value=item.get("requirement", ""))
        main.cell(row=row, column=7, value=item.get("finding", ""))

        rule_status = item.get("rule_status", "")
        main.cell(row=row, column=8, value=RULE_STATUS_LABEL.get(rule_status, rule_status))
        row += 1

    last_body_row = row - 1
    style_body(main, first_row, last_body_row, 8)

    # 摘要列
    counts = {"pass": 0, "fail": 0, "manual": 0, "na": 0}
    for item in results.get("items", []):
        s = item.get("status", "manual")
        if s in counts:
            counts[s] += 1
    total = len(results.get("items", []))

    summary_row = last_body_row + 1
    main.cell(row=summary_row, column=1, value="摘要").font = Font(bold=True)
    main.merge_cells(start_row=summary_row, start_column=1, end_row=summary_row, end_column=2)
    main.cell(row=summary_row, column=3,
              value=f"共 {total} 項：☑ 符合 {counts['pass']}｜☒ 不符合 {counts['fail']}｜"
                    f"⚪ 需人工判讀 {counts['manual']}｜— 不適用 {counts['na']}")
    main.merge_cells(start_row=summary_row, start_column=3, end_row=summary_row, end_column=8)
    for column in range(1, 9):
        cell = main.cell(row=summary_row, column=column)
        cell.border = BORDER
        cell.fill = TOTAL_FILL
        cell.font = Font(bold=True)

    set_widths(main, [6, 16, 10, 10, 22, 40, 55, 20])
    main.freeze_panes = "A5"

    # ---- 分頁 2：摘要統計 ----
    summary = workbook.create_sheet("摘要統計")
    summary["A1"] = "檢核摘要統計"
    summary["A1"].font = TITLE_FONT
    summary.merge_cells("A1:D1")

    write_header(summary, 3, ["狀態", "數量", "百分比", "說明"])

    status_rows = [
        ("☑ 符合", counts["pass"], "符合法條要求"),
        ("☒ 不符合", counts["fail"], "不符合法條要求，需改善"),
        ("⚪ 需人工判讀", counts["manual"], "資料不足或規則未入庫，需人工確認"),
        ("— 不適用", counts["na"], "該法條不適用於本案件"),
    ]
    for offset, (label, count, desc) in enumerate(status_rows):
        r = 4 + offset
        summary.cell(row=r, column=1, value=label)
        summary.cell(row=r, column=2, value=count)
        pct = f"{count / total * 100:.1f}%" if total > 0 else "0.0%"
        summary.cell(row=r, column=3, value=pct)
        summary.cell(row=r, column=4, value=desc)

    style_body(summary, 4, 7, 4)

    # 規則未入庫條號
    not_coded = results.get("not_coded_articles", [])
    nc_row = 9
    summary.cell(row=nc_row, column=1, value="規則未入庫條號").font = HEADER_FONT
    summary.merge_cells(start_row=nc_row, start_column=1, end_row=nc_row, end_column=4)
    summary.cell(row=nc_row, column=1).fill = HEADER_FILL

    if not_coded:
        summary.cell(row=nc_row + 1, column=1, value="、".join(not_coded))
        summary.cell(row=nc_row + 1, column=1).alignment = WRAP_TOP
        summary.merge_cells(start_row=nc_row + 1, start_column=1, end_row=nc_row + 1, end_column=4)
    else:
        summary.cell(row=nc_row + 1, column=1, value="（無）")
        summary.merge_cells(start_row=nc_row + 1, start_column=1, end_row=nc_row + 1, end_column=4)

    # 未核定參數統計
    unverified_row = nc_row + 3
    summary.cell(row=unverified_row, column=1, value="未核定參數統計").font = HEADER_FONT
    summary.merge_cells(start_row=unverified_row, start_column=1, end_row=unverified_row, end_column=4)
    summary.cell(row=unverified_row, column=1).fill = HEADER_FILL

    total_items = len(results.get("items", []))
    unverified_items = [
        i for i in results.get("items", [])
        if "本參數尚未逐條確認" in str(i.get("finding", ""))
    ]
    unverified_count = len(unverified_items)
    ratio = f"{unverified_count / total_items * 100:.1f}%" if total_items > 0 else "0.0%"
    summary.cell(row=unverified_row + 1, column=1,
                 value=f"含未核定參數的檢核項：{unverified_count} / {total_items}（{ratio}）")
    summary.merge_cells(start_row=unverified_row + 1, start_column=1,
                        end_row=unverified_row + 1, end_column=4)

    set_widths(summary, [20, 12, 12, 48])

    return workbook


# ---------------------------------------------------------------------- CLI

def main(argv=None):
    parser = argparse.ArgumentParser(description="法條檢核清單 Excel 產生")
    parser.add_argument("--results", required=True, help="check_results.json 路徑")
    parser.add_argument("--output", help="輸出 xlsx 路徑（預設與 results 同目錄）")
    args = parser.parse_args(argv)

    with open(args.results, encoding="utf-8") as f:
        results = json.load(f)

    workbook = build_checklist(results)
    out = args.output
    if not out:
        out = os.path.join(os.path.dirname(args.results),
                           f"{results.get('case_name', '案件')}-法條檢核清單.xlsx")
    workbook.save(out)

    counts = {}
    for item in results.get("items", []):
        s = item.get("status", "manual")
        counts[s] = counts.get(s, 0) + 1
    total = len(results.get("items", []))
    print(f"✅ 已輸出法條檢核清單：{out}（{total} 項；"
          f"符合 {counts.get('pass', 0)} / 不符合 {counts.get('fail', 0)} / "
          f"需人工判讀 {counts.get('manual', 0)} / 不適用 {counts.get('na', 0)}；2 分頁）")
    return 0


if __name__ == "__main__":
    sys.exit(main())
