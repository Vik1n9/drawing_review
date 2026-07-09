#!/usr/bin/env python3
"""消防安全設備設置標準 Excel 檢核表 HTML 產生器.

將消防人員提供的第 14~31 條判斷用 Excel 表轉為固定版面 HTML，
並依答案 JSON 在對應檢核格內加上紅色勾號。
"""

import argparse
import html
import json
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Optional

try:
    import openpyxl
except ImportError as exc:  # pragma: no cover - exercised by runtime environment
    raise SystemExit("缺少 openpyxl，請先安裝 requirements.txt") from exc


ARTICLE_RE = re.compile(r"^\s*(\d+(?:-\d+)?)\s*條")


class ChecklistInputError(ValueError):
    """輸入檔或答案格式不符合檢核表需求."""


@dataclass(frozen=True)
class ChecklistItem:
    item_id: str
    article: str
    item_no: int
    row: int
    check_cell: str
    text: str


@dataclass(frozen=True)
class CellView:
    row: int
    col: int
    coordinate: str
    value: str
    classes: tuple[str, ...]
    style: str
    answer_id: Optional[str] = None


@dataclass(frozen=True)
class Checklist:
    sheet_name: str
    max_row: int
    max_col: int
    col_widths: tuple[float, ...]
    row_heights: tuple[float, ...]
    rows: tuple[tuple[CellView, ...], ...]
    items: tuple[ChecklistItem, ...]


def css_color(color):
    if color is None:
        return None
    if color.type == "rgb" and color.rgb:
        rgb = color.rgb[-6:]
        return f"#{rgb}"
    if color.type == "indexed":
        indexed = {
            0: "#000000",
            1: "#FFFFFF",
            2: "#FF0000",
            3: "#00FF00",
            4: "#0000FF",
            5: "#FFFF00",
            6: "#FF00FF",
            7: "#00FFFF",
        }
        return indexed.get(color.indexed)
    return None


def excel_width_to_px(width):
    if width is None:
        return 64
    return max(18, round(float(width) * 7 + 5))


def points_to_px(points):
    if points is None:
        return 23
    return max(18, round(float(points) * 4 / 3))


def style_for_cell(cell):
    rules = []
    fill = cell.fill
    if fill and fill.fill_type == "solid":
        fill_color = css_color(fill.fgColor)
        if fill_color:
            rules.append(f"background-color: {fill_color}")

    font = cell.font
    if font:
        if font.name:
            rules.append(f"font-family: {html.escape(font.name)}, 'Noto Sans TC', 'Microsoft JhengHei', sans-serif")
        if font.sz:
            rules.append(f"font-size: {float(font.sz):g}pt")
        if font.bold:
            rules.append("font-weight: 700")
        color = css_color(font.color)
        if color:
            rules.append(f"color: {color}")

    alignment = cell.alignment
    if alignment:
        if alignment.horizontal:
            rules.append(f"text-align: {alignment.horizontal}")
        if alignment.vertical:
            rules.append(f"vertical-align: {alignment.vertical}")
        if alignment.wrap_text:
            rules.append("white-space: pre-wrap")
        else:
            rules.append("white-space: nowrap")

    return "; ".join(rules)


def display_value(formula_cell, data_cell):
    value = formula_cell.value
    if isinstance(value, str) and value.startswith("="):
        value = data_cell.value
    if value is None:
        return ""
    return str(value)


def extract_checklist(workbook_path):
    workbook_path = Path(workbook_path)
    if not workbook_path.exists():
        raise ChecklistInputError(f"找不到 Excel 檢核表: {workbook_path}")

    formula_wb = openpyxl.load_workbook(workbook_path, data_only=False)
    data_wb = openpyxl.load_workbook(workbook_path, data_only=True)
    sheet = formula_wb.active
    data_sheet = data_wb[sheet.title]

    current_article = None
    item_counts = {}
    item_by_row = {}
    items = []

    for row_idx in range(1, sheet.max_row + 1):
        header_value = sheet.cell(row_idx, 1).value
        if isinstance(header_value, str):
            match = ARTICLE_RE.match(header_value)
            if match:
                current_article = match.group(1)
                item_counts.setdefault(current_article, 0)

        text_value = sheet.cell(row_idx, 2).value
        if current_article and text_value not in (None, ""):
            item_counts[current_article] += 1
            item_no = item_counts[current_article]
            item_id = f"{current_article}-{item_no}"
            item = ChecklistItem(
                item_id=item_id,
                article=current_article,
                item_no=item_no,
                row=row_idx,
                check_cell=f"A{row_idx}",
                text=str(text_value),
            )
            items.append(item)
            item_by_row[row_idx] = item

    col_widths = []
    for col_idx in range(1, sheet.max_column + 1):
        letter = openpyxl.utils.get_column_letter(col_idx)
        col_widths.append(excel_width_to_px(sheet.column_dimensions[letter].width))

    row_heights = []
    rows = []
    for row_idx in range(1, sheet.max_row + 1):
        row_heights.append(points_to_px(sheet.row_dimensions[row_idx].height))
        cells = []
        for col_idx in range(1, sheet.max_column + 1):
            cell = sheet.cell(row_idx, col_idx)
            classes = []
            if col_idx == 1 and isinstance(cell.value, str) and ARTICLE_RE.match(cell.value):
                classes.append("article-cell")
            if col_idx == 1 and row_idx in item_by_row:
                classes.append("check-cell")
            if col_idx == 2 and row_idx in item_by_row:
                classes.append("item-text")
            answer_id = item_by_row[row_idx].item_id if col_idx == 1 and row_idx in item_by_row else None
            cells.append(
                CellView(
                    row=row_idx,
                    col=col_idx,
                    coordinate=cell.coordinate,
                    value=display_value(cell, data_sheet.cell(row_idx, col_idx)),
                    classes=tuple(classes),
                    style=style_for_cell(cell),
                    answer_id=answer_id,
                )
            )
        rows.append(tuple(cells))

    return Checklist(
        sheet_name=sheet.title,
        max_row=sheet.max_row,
        max_col=sheet.max_column,
        col_widths=tuple(col_widths),
        row_heights=tuple(row_heights),
        rows=tuple(rows),
        items=tuple(items),
    )


def checked_ids_from_payload(payload):
    if isinstance(payload, list):
        return {str(item) for item in payload}
    if not isinstance(payload, dict):
        raise ChecklistInputError("答案 JSON 必須是物件，或是檢核項目 ID 陣列")

    if "checked" in payload:
        checked = payload["checked"]
        if not isinstance(checked, list):
            raise ChecklistInputError("checked 必須是檢核項目 ID 陣列")
        return {str(item) for item in checked}

    answers = payload.get("answers", [])
    if not isinstance(answers, list):
        raise ChecklistInputError("answers 必須是陣列")

    checked = set()
    for answer in answers:
        if isinstance(answer, str):
            checked.add(answer)
            continue
        if not isinstance(answer, dict):
            raise ChecklistInputError("answers 內每一筆必須是字串 ID 或物件")
        item_id = answer.get("id") or answer.get("item_id")
        if not item_id:
            raise ChecklistInputError("answers 物件缺少 id")
        if answer.get("checked", True):
            checked.add(str(item_id))
    return checked


def validate_checked_ids(checked_ids, checklist):
    valid = {item.item_id for item in checklist.items}
    unknown = sorted(set(checked_ids) - valid)
    if unknown:
        raise ChecklistInputError(f"未知的檢核項目 ID: {', '.join(unknown)}")


def load_answers(answers_path, checklist):
    with open(answers_path, encoding="utf-8") as handle:
        payload = json.load(handle)
    checked_ids = checked_ids_from_payload(payload)
    validate_checked_ids(checked_ids, checklist)
    return checked_ids


def item_manifest(checklist):
    return [
        {
            "id": item.item_id,
            "article": item.article,
            "item_no": item.item_no,
            "row": item.row,
            "check_cell": item.check_cell,
            "text": item.text,
        }
        for item in checklist.items
    ]


def render_answer_template(checklist):
    return json.dumps(
        {
            "checked": [],
            "items": item_manifest(checklist),
        },
        ensure_ascii=False,
        indent=2,
    ) + "\n"


def render_cell(cell, checked_ids):
    class_attr = f' class="{" ".join(cell.classes)}"' if cell.classes else ""
    style_attr = f' style="{html.escape(cell.style)}"' if cell.style else ""
    data_attrs = f' data-cell="{cell.coordinate}"'
    content = html.escape(cell.value)
    if cell.answer_id:
        data_attrs += f' data-answer-id="{html.escape(cell.answer_id)}"'
        if cell.answer_id in checked_ids:
            content = '<span class="red-check">✓</span>'
    return f"<td{class_attr}{style_attr}{data_attrs}>{content}</td>"


def render_html(checklist, checked_ids=None, *, case_name=""):
    checked_ids = set(checked_ids or [])
    validate_checked_ids(checked_ids, checklist)
    title = case_name or checklist.sheet_name
    colgroup = "\n".join(f'<col style="width: {width}px">' for width in checklist.col_widths)
    body_rows = []
    for index, row in enumerate(checklist.rows):
        height = checklist.row_heights[index]
        cells = "".join(render_cell(cell, checked_ids) for cell in row)
        body_rows.append(f'<tr style="height: {height}px">{cells}</tr>')

    return f"""<!DOCTYPE html>
<html lang="zh-Hant">
<head>
<meta charset="utf-8">
<title>{html.escape(title)} — 消防安全設備設置標準檢核表</title>
<style>
body {{
  margin: 24px;
  color: #111;
  background: #fff;
  font-family: "標楷體", "Noto Sans TC", "Microsoft JhengHei", sans-serif;
}}
.sheet-wrap {{
  overflow-x: auto;
}}
table.standard-checklist {{
  border-collapse: collapse;
  table-layout: fixed;
  width: max-content;
}}
.standard-checklist td {{
  border: 1px solid #d0d0d0;
  padding: 2px 4px;
  box-sizing: border-box;
  line-height: 1.25;
  vertical-align: middle;
}}
.article-cell {{
  font-weight: 700;
}}
.check-cell {{
  text-align: center;
  vertical-align: middle;
}}
.item-text {{
  white-space: pre-wrap;
}}
.red-check {{
  color: #d00000;
  font-family: "Noto Sans TC", "Microsoft JhengHei", sans-serif;
  font-size: 24px;
  font-weight: 800;
  line-height: 1;
}}
.meta {{
  margin: 0 0 12px;
  color: #666;
  font-family: "Noto Sans TC", "Microsoft JhengHei", sans-serif;
  font-size: 13px;
}}
@media print {{
  body {{ margin: 8mm; }}
  .sheet-wrap {{ overflow: visible; }}
  .standard-checklist td {{ -webkit-print-color-adjust: exact; print-color-adjust: exact; }}
}}
</style>
</head>
<body>
<div class="meta">紅色勾號由答案 JSON 產生；未列入答案的檢核格保持空白。</div>
<div class="sheet-wrap">
<table class="standard-checklist" aria-label="{html.escape(checklist.sheet_name)}">
<colgroup>
{colgroup}
</colgroup>
<tbody>
{chr(10).join(body_rows)}
</tbody>
</table>
</div>
</body>
</html>
"""


def default_output_path(input_path):
    return Path("rules") / "checklists" / f"{Path(input_path).stem}.html"


def main(argv=None):
    parser = argparse.ArgumentParser(description="由第 14~31 條 Excel 標準表產生最終檢核 HTML")
    parser.add_argument("--input", required=True, help="消防標準表 Excel 路徑")
    parser.add_argument("--answers", help="答案 JSON；可用 checked 或 answers 指定檢核項目 ID")
    parser.add_argument("--output", help="輸出 HTML 路徑；預設寫入 rules/checklists/{Excel檔名}.html")
    parser.add_argument("--case-name", default="", help="HTML title 使用的案件名稱")
    parser.add_argument("--dump-answer-template", help="輸出可填寫的答案範本 JSON 路徑")
    args = parser.parse_args(argv)

    checklist = extract_checklist(args.input)

    if args.dump_answer_template:
        template_path = Path(args.dump_answer_template)
        template_path.parent.mkdir(parents=True, exist_ok=True)
        template_path.write_text(render_answer_template(checklist), encoding="utf-8")
        print(f"已輸出答案範本：{template_path}")

    checked_ids = load_answers(args.answers, checklist) if args.answers else set()
    output_path = Path(args.output) if args.output else default_output_path(args.input)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(render_html(checklist, checked_ids, case_name=args.case_name), encoding="utf-8")
    print(f"已輸出檢核 HTML：{output_path}（紅勾 {len(checked_ids)} 項）")


if __name__ == "__main__":
    try:
        main()
    except ChecklistInputError as exc:
        raise SystemExit(f"輸入錯誤：{exc}") from exc
