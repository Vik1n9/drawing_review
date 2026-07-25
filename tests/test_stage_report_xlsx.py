import contextlib
import io
import json
import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook

from tools.stage_report_xlsx import (
    MANUAL,
    DEFAULT_LAW_PATH,
    add_references,
    article_18_fallback,
    build_first_stage,
    build_stage_two,
    ensure_article_18,
    format_basis,
    main,
    normalize_law_rows,
    read_law_rows,
    scope_only_note,
)


def sample_case():
    return {
        "case_name": "測試案",
        "regulation_version": "測試版",
        "use_permit": {"permit_no": {"value": "測使字第001號", "confidence": "high", "source": "manual"},
                       "total_floor_area": {"value": 900, "confidence": "high", "source": "manual"}},
        "interior_renovation": {
            "floors": {"value": ["3F"], "confidence": "high", "source": "manual"},
            "area": {"value": 350, "confidence": "high", "source": "manual"},
            "post_renovation_use": {"value": "乙7", "confidence": "high", "source": "manual"},
        },
        "floors": [
            {"floor": "B1", "area": {"value": 300, "confidence": "high"},
             "floor_position": {"value": "basement", "confidence": "high"},
             "use_category": {"value": "丙1", "label": "停車空間", "legal_basis": "§12 第3款第1目",
                              "confidence": "high"}},
            {"floor": "3F", "area": {"value": 350, "confidence": "high"},
             "floor_position": {"value": "ordinary", "confidence": "high"},
             "windowless": {"value": True, "confidence": "high"},
             "use_category": {"value": "乙7", "label": "集合住宅", "legal_basis": "§12 第2款第7目",
                              "confidence": "high"},
             "rooms": [{"name": "臥室", "area": 20, "ceiling_height": 3.1}]},
            {"floor": "4F", "area": {"value": 250, "confidence": "high"},
             "floor_position": {"value": "ordinary", "confidence": "low"},
             "use_category": {"value": "乙6", "label": "辦公室", "legal_basis": "§12 第2款第6目",
                              "confidence": "high"}},
        ],
        "building": {"mixed_use_assessment": {"is_mixed_use": True, "category_candidate": "戊1",
                                              "confidence": "high", "source": "manual"}},
        "manual_review_items": [],
    }


class FirstStageWorkbookTest(unittest.TestCase):
    def setUp(self):
        self.workbook = build_first_stage(sample_case())
        self.main = self.workbook["複合用途檢討"]

    def test_sheet_structure_matches_template(self):
        self.assertEqual(self.workbook.sheetnames,
                         ["複合用途檢討", "用途判斷依據", "來源資料", "各樓層高度"])

    def test_column_order_is_fixed(self):
        headers = [self.main.cell(row=4, column=c).value for c in range(1, 8)]
        self.assertEqual(headers, ["樓層", "各層用途", "樓地板面積(m²)", "各層變更後用途",
                                   "本次申請範圍樓地板面積(m²)", "樓層屬性", "備註"])

    def test_exactly_three_blank_rows_then_total(self):
        # 3 層資料 → 列 5,6,7；空白 8,9,10；合計 11
        self.assertEqual(self.main.cell(row=5, column=1).value, "B1")
        self.assertEqual(self.main.cell(row=7, column=1).value, "4F")
        for row in (8, 9, 10):
            self.assertIsNone(self.main.cell(row=row, column=1).value)
        self.assertEqual(self.main.cell(row=11, column=1).value, "樓地板面積合計")
        self.assertEqual(self.main.cell(row=11, column=3).value, "=SUM(C5:C10)")

    def test_judgment_section_follows_one_blank_spacer(self):
        self.assertIsNone(self.main.cell(row=12, column=1).value)
        self.assertEqual(self.main.cell(row=13, column=1).value, "複合用途建築物判定")
        self.assertEqual(self.main.cell(row=14, column=1).value, "第12條用途分類")

    def test_each_classification_gets_its_own_row_and_judgment_appears_once(self):
        first = 15
        classifications = [self.main.cell(row=first + i, column=1).value for i in range(3)]
        self.assertEqual(len(classifications), len(set(classifications)))
        self.assertTrue(all("§12" in str(c) for c in classifications))
        self.assertEqual(self.main.cell(row=first, column=5).value, "戊1複合用途建築物")
        self.assertIsNone(self.main.cell(row=first + 1, column=5).value)

    def test_classification_without_legal_basis_is_flagged_not_silently_shortened(self):
        case = sample_case()
        del case["floors"][0]["use_category"]["legal_basis"]
        sheet = build_first_stage(case)["複合用途檢討"]
        flagged = sheet.cell(row=15, column=1).value
        self.assertIn("丙1", flagged)
        self.assertIn(MANUAL, flagged)
        self.assertIn("§12 款目未載", flagged)

    def test_floor_attribute_blank_when_evidence_insufficient(self):
        self.assertEqual(self.main.cell(row=5, column=6).value, "地下層")
        self.assertEqual(self.main.cell(row=6, column=6).value, "無開口樓層")
        self.assertEqual(self.main.cell(row=7, column=6).value, "")

    def test_application_scope_only_marked_on_renovation_floor(self):
        self.assertIsNone(self.main.cell(row=5, column=7).value)
        self.assertEqual(self.main.cell(row=6, column=7).value, "本次申請範圍")
        self.assertEqual(self.main.cell(row=6, column=5).value, 350)
        self.assertIsNone(self.main.cell(row=7, column=7).value)

    def test_dropdown_validations_keep_blank_option(self):
        formulas = [dv.formula1 for dv in self.main.data_validations.dataValidation]
        self.assertIn('"地下層,無開口樓層,屋頂層,一般樓層,"', formulas)
        self.assertIn('"本次申請範圍,"', formulas)
        self.assertIn('"單一用途建築物,戊1複合用途建築物,戊2複合用途建築物,"', formulas)

    def test_unconfirmed_judgment_leaves_verdict_blank(self):
        case = sample_case()
        case["building"]["mixed_use_assessment"]["source"] = "derived"
        sheet = build_first_stage(case)["複合用途檢討"]
        self.assertEqual(sheet.cell(row=15, column=5).value, "")

    def test_floor_heights_cover_every_floor_in_order(self):
        heights = self.workbook["各樓層高度"]
        self.assertEqual([heights.cell(row=r, column=1).value for r in (4, 5, 6)], ["B1", "3F", "4F"])
        self.assertEqual(heights.cell(row=5, column=2).value, 3.1)
        self.assertIsNone(heights.cell(row=4, column=2).value)
        self.assertEqual(heights.cell(row=4, column=4).value, "未提供")


class StageTwoWorkbookTest(unittest.TestCase):
    def law_rows(self):
        return [
            ("17條下列場所應設置自動撒水設備", None),
            (None, "一. 甲類場所."),
            ("19條下列場所應設置火警自動警報設備", None),
            (None, "一. 甲類場所."),
            (None, "二. 六層以上十層以下之樓層."),
        ]

    def test_article_18_inserted_between_17_and_19(self):
        rows = ensure_article_18(self.law_rows())
        titles = [t for t, _ in rows if t]
        self.assertEqual(titles[0][:3], "17條")
        self.assertTrue(titles[1].startswith("18條"))
        self.assertTrue(titles[-1].startswith("19條"))

    def test_article_18_text_has_a_single_source_of_truth(self):
        """§18 條文轉錄只存在 rules/article18_equipment_options.json 一份。

        工具內不得再有第二份複本——多處複本會隨修法漂移
        （rules/stage_two_judgment_rules.md「維護規則」）。
        """
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "opts.json"
            path.write_text(json.dumps({
                "article_title": "測試標題",
                "items": [{"item": 1, "condition": "測試款項條件。", "options": ["泡沫"]}],
            }, ensure_ascii=False), encoding="utf-8")
            rows = article_18_fallback(str(path))
        self.assertEqual(rows, [("18條測試標題", None), (None, "一. 測試款項條件。")])

    def test_article_18_fallback_degrades_when_rules_file_missing(self):
        self.assertEqual(article_18_fallback("/nonexistent.json"), [])
        rows = self.law_rows()
        self.assertEqual(ensure_article_18(rows, "/nonexistent.json"), rows)

    def test_repo_law_table_already_contains_article_18(self):
        rows = read_law_rows(DEFAULT_LAW_PATH)
        self.assertTrue(any(t and t.startswith("18條") for t, _ in rows),
                        "rules/checklists 的法條判斷表應已含第 18 條")
        self.assertEqual(rows, ensure_article_18(rows), "已含 §18 時不應重複插入")

    def test_references_skip_title_only_rows(self):
        rows = add_references(self.law_rows())
        self.assertIsNone(rows[0]["reference"])
        self.assertEqual(rows[1]["reference"], "17-1")
        self.assertEqual(rows[4]["reference"], "19-2")

    def test_checked_rows_marked_and_notes_limited_to_scope(self):
        decisions = {
            "checkedRefs": ["17-1", "19-2"],
            "notesByRef": {"17-1": "停車空間非本次申請範圍", "19-2": "六層以上建築物適用"},
        }
        workbook, _ = build_stage_two(self.law_rows(), decisions)
        sheet = workbook["消防安全設備設置標準"]
        self.assertEqual(sheet.cell(row=3, column=1).value, "✓")   # 17-1
        self.assertEqual(sheet.cell(row=3, column=4).value, "非本次申請範圍")
        # 19-2 為最末列（§18 補入 7 款後）：已勾選，但判定說明不得留在備註欄
        self.assertEqual(sheet.cell(row=14, column=1).value, "✓")
        self.assertEqual(sheet.cell(row=14, column=4).value, "")

    def test_article_18_shows_selectable_equipment_not_foam_only(self):
        workbook, rows = build_stage_two(self.law_rows(), {})
        sheet = workbook["消防安全設備設置標準"]
        titles = [sheet.cell(row=r, column=2).value for r in range(2, sheet.max_row + 1)]
        self.assertIn("18條下列場所應設置自動滅火設備", titles)
        self.assertIn("泡沫、乾粉", titles)
        self.assertIn("二氧化碳或惰性氣體、鹵化烴、乾粉；或預動式自動撒水設備", titles)
        # 第 8、9 款不是設備觸發條件，不列於表中
        self.assertFalse(any(r["article"] == "18" and (r["item"] or 0) > 7 for r in rows))

    def test_equipment_sheet_derived_from_checked_refs(self):
        workbook, _ = build_stage_two(self.law_rows(), {"checkedRefs": ["19-1", "19-2", "17-1"]})
        sheet = workbook["應設置設備"]
        rows = {sheet.cell(row=r, column=1).value: sheet.cell(row=r, column=2).value
                for r in range(2, sheet.max_row + 1)}
        self.assertEqual(rows["自動撒水設備"], "第17條第1款")
        self.assertEqual(rows["火警自動警報設備"], "第19條第1、2款")

    def test_pending_sheet_defaults_to_none_row(self):
        workbook, _ = build_stage_two(self.law_rows(), {})
        sheet = workbook["待釐清事項"]
        self.assertEqual(sheet.cell(row=2, column=1).value, "無")

    def test_sheet_order_and_headers(self):
        workbook, _ = build_stage_two(self.law_rows(), {})
        self.assertEqual(workbook.sheetnames, ["消防安全設備設置標準", "應設置設備", "待釐清事項"])
        main = workbook["消防安全設備設置標準"]
        self.assertEqual([main.cell(row=1, column=c).value for c in range(1, 5)],
                         ["判斷", "條文/設備", "款項條件", "備註"])

    def test_format_basis_and_scope_note_helpers(self):
        # 依條號排序，款次併列
        self.assertEqual(format_basis(["24-4", "24-3", "22"]), "第22條、第24條第3、4款")
        self.assertEqual(scope_only_note("停車空間非本次申請範圍"), "非本次申請範圍")
        self.assertEqual(scope_only_note("因火警自動警報設備符合"), "")

    def test_normalize_law_rows_accepts_output_header_format(self):
        values = [("判斷", "條文/設備", "款項條件", "備註"), ("", "14條…", "", "")]
        self.assertEqual(normalize_law_rows(values), [("14條…", None)])


class CliTest(unittest.TestCase):
    def test_both_subcommands_write_workbooks(self):
        with tempfile.TemporaryDirectory() as tmp:
            case_path = Path(tmp) / "case.json"
            case_path.write_text(json.dumps(sample_case(), ensure_ascii=False), encoding="utf-8")

            with contextlib.redirect_stdout(io.StringIO()):
                self.assertEqual(main(["first-stage", "--case", str(case_path)]), 0)
            first = Path(tmp) / "測試案-第一階段-複合用途及樓層屬性檢討.xlsx"
            self.assertTrue(first.exists())
            self.assertIn("複合用途檢討", load_workbook(first).sheetnames)

            decisions = Path(tmp) / "stage2_decisions.json"
            decisions.write_text(json.dumps({"checkedRefs": ["14-1"]}), encoding="utf-8")
            with contextlib.redirect_stdout(io.StringIO()):
                self.assertEqual(main(["stage-two", "--law", DEFAULT_LAW_PATH,
                                       "--decisions", str(decisions), "--case", str(case_path)]), 0)
            second = Path(tmp) / "測試案-第二階段-設置標準檢討.xlsx"
            self.assertTrue(second.exists())
            self.assertIn("待釐清事項", load_workbook(second).sheetnames)


if __name__ == "__main__":
    unittest.main()
